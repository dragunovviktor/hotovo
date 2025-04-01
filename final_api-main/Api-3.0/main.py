from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Request
from pydantic import BaseModel, validator,AnyUrl, Field
from typing import List, Optional
from datetime import datetime
import os
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, ForeignKey, Text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
import uuid
from typing import List, Optional, Union
from fastapi.templating import Jinja2Templates
from dateutil.parser import parse
from fastapi import UploadFile, File, HTTPException, Form, status
from fastapi.responses import FileResponse
from pathlib import Path
from fastapi import Response
from sqlalchemy.orm import relationship  # Добавьте этот импорт
import aiofiles
import uuid
from openpyxl import load_workbook  # Для работы с картинками в xlsx
import pandas as pd
import openpyxl
from io import BytesIO
from sqlalchemy import text
from PIL import Image
import io

DATABASE_URL = "postgresql://sber_admin:securepass@db:5432/sber_api"
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base = declarative_base()
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Модели базы данных
class Branch(Base):
    __tablename__ = "branches"
    id = Column(Integer, primary_key=True, index=True)
    address = Column(String, index=True)
    internal_code = Column(String, unique=True)
    latitude = Column(String)
    longitude = Column(String)


class ObjectType(Base):
    __tablename__ = "object_types"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String)  # газон, крыльцо, тротуар
    measure_unit = Column(String)  # м², пог. м


class BranchObject(Base):
    __tablename__ = "branch_objects"
    id = Column(Integer, primary_key=True, index=True)
    branch_id = Column(Integer, ForeignKey("branches.id"))
    object_type_id = Column(Integer, ForeignKey("object_types.id"))
    name = Column(String)
    area = Column(String)
    description = Column(String)
    zone = Column(String)
    surface_type = Column(String)
    work_period = Column(String)
    responsible_person = Column(String, nullable=True)
    work_type = Column(String, nullable=True)
    measure_unit = Column(String, nullable=True)


class MaintenancePlan(Base):
    __tablename__ = "maintenance_plans"
    id = Column(Integer, primary_key=True, index=True)
    branch_id = Column(Integer, ForeignKey("branches.id"))
    object_id = Column(Integer, ForeignKey("branch_objects.id"), nullable=True)
    work_type = Column(String)
    frequency = Column(String)
    next_maintenance_date = Column(DateTime)


class BranchTerritory(Base):
    __tablename__ = "branch_territories"
    id = Column(Integer, primary_key=True, index=True)
    branch_id = Column(Integer, ForeignKey("branches.id"))
    responsible_person = Column(String)
    legal_area = Column(Float)
    summer_area = Column(Float)
    winter_area = Column(Float)
    municipal_area = Column(Float)
    client_area = Column(Float)
    scheme_attachment_id = Column(Integer, ForeignKey("branch_attachments.id"), nullable=True)

    scheme = relationship("BranchAttachment", foreign_keys=[scheme_attachment_id])

class CompletedWork(Base):
    __tablename__ = "completed_works"
    id = Column(Integer, primary_key=True, index=True)
    branch_id = Column(Integer, ForeignKey("branches.id"))
    object_id = Column(Integer, ForeignKey("branch_objects.id"), nullable=True)
    work_type = Column(String)
    completion_date = Column(DateTime)
    responsible_person = Column(String)
    notes = Column(Text)


class BranchAttachment(Base):
    __tablename__ = "branch_attachments"
    id = Column(Integer, primary_key=True, index=True)
    branch_id = Column(Integer, ForeignKey("branches.id"))
    object_id = Column(Integer, ForeignKey("branch_objects.id"), nullable=True)
    file_type = Column(String)
    file_url = Column(String)
    original_filename = Column(String)
    uploaded_at = Column(DateTime, default=datetime.utcnow)


Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="Sber Branches API",
    description="API для управления территориями отделений Сбербанка",
    version="1.0.0",
    docs_url="/docs",  # Измените на /docs вместо /api/docs
    redoc_url="/redoc"  # Измените на /redoc вместо /api/redoc
)

# Pydantic модели

class BranchBase(BaseModel):
    address: str
    internal_code: str


class BranchCreate(BaseModel):
    address: str
    internal_code: str
    latitude: Optional[Union[float, str]] = None  # Принимаем и float, и string
    longitude: Optional[Union[float, str]] = None

class BranchCreate(BranchBase):
    latitude: Optional[float] = None
    longitude: Optional[float] = None

    @validator('latitude', 'longitude')
    def validate_coordinates(cls, v):
        if v is not None:
            try:
                float(v)
            except ValueError:
                raise ValueError("Координата должна быть числом")
        return v

class BranchResponse(BranchBase):
    latitude: Optional[str] = None
    longitude: Optional[str] = None

    class Config:
        orm_mode = True


class AttachmentResponse(BaseModel):
    id: int
    branch_id: Optional[int] = None  # Make it optional
    branch_code: Optional[str] = None  # Add branch_code if needed
    object_id: Optional[int]
    file_type: str
    file_url: str
    original_filename: str
    uploaded_at: datetime

    class Config:
        orm_mode = True

class AttachmentCreate(BaseModel):
    branch_code: str
    object_id: Optional[int] = None
    file_type: str = "photo"
    file_url: AnyUrl  # Автоматическая валидация URL


class ObjectCreate(BaseModel):
    branch_code: str
    object_type_id: int
    name: str
    area: str
    description: Optional[str] = None

class ObjectResponse(BaseModel):
    id: int
    branch_code: str
    address: str
    responsible: str
    zone: str
    surface_type: str
    work_type: str
    measure_unit: str
    work_volume: str
    work_period: str

    class Config:
        orm_mode = True

class MaintenancePlanCreate(BaseModel):
    branch_code: str
    object_id: Optional[int] = None
    work_type: str
    frequency: str
    next_maintenance_date: str  # Принимаем строку

    @validator('next_maintenance_date')
    def parse_date(cls, v):
        try:
            return parse(v)
        except ValueError:
            raise ValueError("Некорректный формат даты. Попробуйте 'YYYY-MM-DD', 'DD.MM.YYYY' или подобные форматы")

class MaintenancePlanResponse(BaseModel):
    branch_code: str
    object_id: Optional[int] = None
    work_type: str
    frequency: str
    next_maintenance_date: datetime  # В ответе возвращаем datetime

    class Config:
        orm_mode = True

class CompletedWorkCreate(BaseModel):
    branch_code: str
    object_id: Optional[int] = None
    work_type: str
    completion_date: str  # Принимаем строку от пользователя
    responsible_person: str
    notes: Optional[str] = None

    @validator('completion_date')
    def parse_date(cls, v):
        try:
            return parse(v)
        except ValueError:
            raise ValueError("Некорректный формат даты. Попробуйте 'YYYY-MM-DD', 'DD.MM.YYYY' или подобные форматы")

class CompletedWorkResponse(BaseModel):
    branch_code: str
    object_id: Optional[int] = None
    work_type: str
    completion_date: datetime  # В ответе возвращаем datetime
    responsible_person: str
    notes: Optional[str] = None

    class Config:
        orm_mode = True

class ObjectTypeCreate(BaseModel):
    name: str
    measure_unit: str

class ObjectTypeResponse(ObjectTypeCreate):
    id: int

    class Config:
        orm_mode = True

# Dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

templates = Jinja2Templates(directory="templates")

Base.metadata.create_all(bind=engine)


def extract_images_from_excel(file_path: str):
    """Извлекает все изображения из Excel файла"""
    wb = openpyxl.load_workbook(filename=file_path)
    images = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for img in ws._images:
            # Получаем изображение
            img_data = img._data()
            # Сохраняем по номеру строки (Excel нумерует строки с 1)
            row = img.anchor._from.row + 1 if hasattr(img, 'anchor') else img.anchor.row + 1
            images[row] = img_data

    return images


@app.get("/api_ui", include_in_schema=False)
async def get_api_info(request: Request):
    return templates.TemplateResponse("api_ui.html", {"request": request})

# Эндпоинты для администраторов
@app.post("/api/branches", response_model=BranchResponse)
def create_branch(branch: BranchCreate, db: Session = Depends(get_db)):
    existing = db.query(Branch).filter(Branch.internal_code == branch.internal_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Филиал с таким кодом уже существует")

    db_branch = Branch(
        address=branch.address,
        internal_code=branch.internal_code,
        latitude=str(branch.latitude) if branch.latitude is not None else None,
        longitude=str(branch.longitude) if branch.longitude is not None else None
    )

    db.add(db_branch)
    db.commit()
    db.refresh(db_branch)
    return db_branch


@app.post("/api/db/migrate")
def run_migrations(db: Session = Depends(get_db)):
    try:
        # Проверяем существование колонки
        check_column = db.execute(text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name='branch_territories' 
            AND column_name='scheme_attachment_id'
        """)).fetchone()

        if not check_column:
            # Добавляем колонку и внешний ключ
            db.execute(text("""
                ALTER TABLE branch_territories 
                ADD COLUMN scheme_attachment_id INTEGER;

                ALTER TABLE branch_territories 
                ADD CONSTRAINT fk_scheme_attachment 
                FOREIGN KEY (scheme_attachment_id) 
                REFERENCES branch_attachments(id);
            """))
            db.commit()
            return {"status": "success", "message": "Миграция выполнена: добавлена колонка scheme_attachment_id"}

        return {"status": "info", "message": "Колонка уже существует"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка миграции: {str(e)}"
        )

@app.get("/api/branches/by-code/{branch_code}/attachments", response_model=List[AttachmentResponse])
def get_branch_attachments(
    branch_code: str,
    object_id: Optional[int] = None,
    file_type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    try:
        # Находим филиал с JOIN к attachments
        branch = db.query(Branch)\
            .filter(Branch.internal_code == branch_code)\
            .first()

        if not branch:
            raise HTTPException(status_code=404, detail="Филиал не найден")

        # Строим запрос с явным указанием полей
        query = db.query(
            BranchAttachment.id,
            BranchAttachment.branch_id,
            Branch.internal_code.label("branch_code"),
            BranchAttachment.object_id,
            BranchAttachment.file_type,
            BranchAttachment.file_url,
            BranchAttachment.uploaded_at
        ).join(Branch)\
         .filter(BranchAttachment.branch_id == branch.id)

        # Применяем фильтры
        if object_id is not None:
            query = query.filter(BranchAttachment.object_id == object_id)

        if file_type is not None:
            query = query.filter(BranchAttachment.file_type == file_type)

        # Выполняем запрос и преобразуем результат
        attachments = query.all()

        return [{
            "id": a.id,
            "branch_id": a.branch_id,
            "branch_code": a.branch_code,
            "object_id": a.object_id,
            "file_type": a.file_type,
            "file_url": a.file_url,
            "uploaded_at": a.uploaded_at
        } for a in attachments]

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при получении вложений: {str(e)}"
        )

@app.put("/api/branches/by-code/{internal_code}", response_model=BranchResponse)
def update_branch_by_code(internal_code: str, branch: BranchBase, db: Session = Depends(get_db)):
    # Находим филиал по internal_code
    db_branch = db.query(Branch).filter(Branch.internal_code == internal_code).first()
    if not db_branch:

        raise HTTPException(status_code=404, detail="Филиал с указанным кодом не найден")

    # Обновляем только переданные поля (исключая internal_code)
    update_data = branch.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_branch, field, value)

    db.commit()
    db.refresh(db_branch)
    return db_branch

@app.post("/api/object-types", response_model=ObjectTypeResponse)
def create_object_type(obj_type: ObjectTypeCreate, db: Session = Depends(get_db)):
    # Проверяем, нет ли уже типа с таким именем
    existing = db.query(ObjectType).filter(ObjectType.name == obj_type.name).first()
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Тип объекта с названием '{obj_type.name}' уже существует"
        )

    db_type = ObjectType(
        name=obj_type.name,
        measure_unit=obj_type.measure_unit
    )
    db.add(db_type)
    db.commit()
    db.refresh(db_type)
    return db_type


@app.put("/api/branches/{branch_id}", response_model=BranchResponse)
def update_branch(branch_id: int, branch: BranchCreate, db: Session = Depends(get_db)):
    db_branch = db.query(Branch).filter(Branch.id == branch_id).first()
    if not db_branch:
        raise HTTPException(status_code=404, detail="Филиал не найден")

    update_data = branch.dict(exclude_unset=True)
    for field, value in update_data.items():
        if field in ['latitude', 'longitude'] and value is not None:
            setattr(db_branch, field, str(value))
        elif value is not None:
            setattr(db_branch, field, value)

    db.commit()
    db.refresh(db_branch)
    return db_branch

@app.get("/", include_in_schema=False)
async def file_upload_page(request: Request):
    return templates.TemplateResponse("apple_style_upload.html", {"request": request})

@app.post("/api/objects", response_model=ObjectCreate)
def create_object(obj: ObjectCreate, db: Session = Depends(get_db)):
    # Находим филиал по internal_code
    branch = db.query(Branch).filter(Branch.internal_code == obj.branch_code).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Филиал с указанным кодом не найден")

    # Проверяем существование типа объекта
    if not db.query(ObjectType).filter(ObjectType.id == obj.object_type_id).first():
        raise HTTPException(status_code=404, detail="Тип объекта не найден")

    db_object = BranchObject(
        branch_id=branch.id,  # Используем найденный ID филиала
        object_type_id=obj.object_type_id,
        name=obj.name,
        area=obj.area,
        description=obj.description
    )

    db.add(db_object)
    db.commit()
    db.refresh(db_object)

    return {
        "branch_code": obj.branch_code,  # Возвращаем переданный код
        "object_type_id": db_object.object_type_id,
        "name": db_object.name,
        "area": db_object.area,
        "description": db_object.description
    }


@app.get("/api/objects", response_model=List[ObjectResponse])
def get_all_objects(db: Session = Depends(get_db)):
    objects = db.query(BranchObject).all()

    result = []
    for obj in objects:
        branch = db.query(Branch).filter(Branch.id == obj.branch_id).first()
        obj_type = db.query(ObjectType).filter(ObjectType.id == obj.object_type_id).first()

        result.append({
            "id": obj.id,
            "branch_code": branch.internal_code if branch else "-",
            "address": branch.address if branch else "-",
            "responsible": obj.responsible_person or "-",
            "zone": obj.zone or "-",
            "surface_type": obj.surface_type or "-",
            "work_type": obj.work_type or "-",
            "measure_unit": obj.measure_unit or (obj_type.measure_unit if obj_type else "-"),
            "work_volume": obj.area or "-",
            "work_period": obj.work_period or "-"
        })

    return result

@app.get("/api_ui_for_users", include_in_schema=False)
async def get_api_info(request: Request):
    return templates.TemplateResponse("api_ui_for_users.html", {"request": request})

@app.post("/api/maintenance", response_model=MaintenancePlanResponse)
def create_maintenance_plan(plan: MaintenancePlanCreate, db: Session = Depends(get_db)):
    # Находим филиал по коду
    branch = db.query(Branch).filter(Branch.internal_code == plan.branch_code).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Филиал с указанным кодом не найден")

    # Проверяем объект, если указан
    if plan.object_id:
        if not db.query(BranchObject).filter(BranchObject.id == plan.object_id).first():
            raise HTTPException(status_code=404, detail="Объект филиала не найден")

    db_plan = MaintenancePlan(
        branch_id=branch.id,
        object_id=plan.object_id,
        work_type=plan.work_type,
        frequency=plan.frequency,
        next_maintenance_date=plan.next_maintenance_date  # Уже преобразовано в datetime
    )

    db.add(db_plan)
    db.commit()
    db.refresh(db_plan)

    return {
        "branch_code": plan.branch_code,
        "object_id": db_plan.object_id,
        "work_type": db_plan.work_type,
        "frequency": db_plan.frequency,
        "next_maintenance_date": db_plan.next_maintenance_date
    }


@app.post("/api/upload/territories")
async def upload_territories_table(
        file: UploadFile = File(...),
        db: Session = Depends(get_db)
):
    try:
        # Сохраняем временный файл
        file_path = f"temp_{file.filename}"
        with open(file_path, "wb") as buffer:
            buffer.write(await file.read())

        # Извлекаем изображения из Excel
        images = extract_images_from_excel(file_path)
        print(f"Найдено изображений: {len(images)}")

        # Читаем данные из Excel
        try:
            df = pd.read_excel(file_path)
            # Проверяем количество столбцов
            if len(df.columns) < 8:
                raise HTTPException(
                    status_code=400,
                    detail="Файл должен содержать 8 столбцов: branch_code, address, responsible, legal_area, summer_area, winter_area, municipal_area, client_area"
                )

            # Берем только первые 8 столбцов
            df = df.iloc[:, :8]
            df.columns = [
                'branch_code', 'address', 'responsible',
                'legal_area', 'summer_area', 'winter_area',
                'municipal_area', 'client_area'
            ]
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Ошибка чтения Excel файла: {str(e)}"
            )

        results = []

        for index, row in df.iterrows():
            branch_code = None
            try:
                branch_code = str(row['branch_code']).strip()
                print(f"Обработка филиала: {branch_code}")

                # Находим или создаем филиал
                branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
                if not branch:
                    branch = Branch(
                        internal_code=branch_code,
                        address=row['address'],
                        latitude=None,
                        longitude=None
                    )
                    db.add(branch)
                    db.commit()
                    print(f"Создан новый филиал: {branch_code}")

                # Находим или создаем территорию филиала
                territory = db.query(BranchTerritory).filter(
                    BranchTerritory.branch_id == branch.id
                ).first()

                if not territory:
                    territory = BranchTerritory(branch_id=branch.id)
                    db.add(territory)
                    print(f"Создана новая территория для филиала: {branch_code}")

                # Обновляем данные территории
                territory.responsible_person = row['responsible']
                territory.legal_area = float(row['legal_area']) if pd.notna(row['legal_area']) else None
                territory.summer_area = float(row['summer_area']) if pd.notna(row['summer_area']) else None
                territory.winter_area = float(row['winter_area']) if pd.notna(row['winter_area']) else None
                territory.municipal_area = float(row['municipal_area']) if pd.notna(row['municipal_area']) else None
                territory.client_area = float(row['client_area']) if pd.notna(row['client_area']) else None

                # Сохраняем схему, если есть для этой строки
                # Excel строки нумеруются с 1, header - строка 1, данные с 2
                excel_row_number = index + 2

                if excel_row_number in images:
                    print(f"Найдена схема для строки {excel_row_number}")

                    # Генерируем уникальное имя файла
                    unique_filename = f"scheme_{branch_code}_{uuid.uuid4()}.png"
                    scheme_path = os.path.join(UPLOAD_DIR, unique_filename)

                    # Сохраняем изображение на диск
                    try:
                        with open(scheme_path, "wb") as img_file:
                            img_file.write(images[excel_row_number])
                        print(f"Схема сохранена: {scheme_path}")
                    except Exception as e:
                        print(f"Ошибка сохранения схемы: {str(e)}")
                        raise

                    # Создаем запись в базе о вложении
                    attachment = BranchAttachment(
                        branch_id=branch.id,
                        file_type="scheme",
                        file_url=unique_filename,
                        original_filename=f"scheme_{branch_code}.png"
                    )
                    db.add(attachment)
                    db.commit()

                    # Привязываем схему к территории
                    territory.scheme_attachment_id = attachment.id
                    print(f"Схема привязана к территории: {attachment.id}")

                db.commit()

                results.append({
                    "branch_code": branch_code,
                    "status": "success",
                    "message": "Данные территории обновлены",
                    "has_scheme": excel_row_number in images
                })

            except Exception as e:
                print(f"Ошибка при обработке филиала {branch_code}: {str(e)}")
                results.append({
                    "branch_code": branch_code if branch_code else "unknown",
                    "status": "error",
                    "error": str(e)
                })
                db.rollback()

        # Удаляем временный файл
        if os.path.exists(file_path):
            os.remove(file_path)

        return {"results": results}

    except Exception as e:
        # Удаляем временный файл, если он был создан
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(
            status_code=400,
            detail=f"Ошибка обработки файла: {str(e)}"
        )

@app.post("/api/upload/objects")
async def upload_objects_table(
        file: UploadFile = File(...),
        db: Session = Depends(get_db)
):
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(await file.read()))
        else:
            df = pd.read_excel(BytesIO(await file.read()))

        df.columns = [
            'branch_code', 'address', 'responsible',
            'zone', 'surface_type', 'work_type',
            'measure_unit', 'work_volume', 'period'
        ]

        results = []

        for _, row in df.iterrows():
            try:
                branch_code = str(row['branch_code']).strip()
                branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
                if not branch:
                    results.append({
                        "branch_code": branch_code,
                        "status": "error",
                        "error": "Филиал не найден"
                    })
                    continue

                object_type_name = f"{row['zone']} - {row['surface_type']}"
                object_type = db.query(ObjectType).filter(
                    ObjectType.name == object_type_name
                ).first()

                if not object_type:
                    object_type = ObjectType(
                        name=object_type_name,
                        measure_unit=row['measure_unit']
                    )
                    db.add(object_type)
                    db.commit()

                # Создаем объект только с существующими полями
                branch_object = BranchObject(
                    branch_id=branch.id,
                    object_type_id=object_type.id,
                    name=object_type_name,
                    area=str(row['work_volume']),
                    description=f"Ответственный: {row['responsible']}. Период: {row['period']}",
                    zone=row['zone'],
                    surface_type=row['surface_type'],
                    work_period=row['period'],
                    # Эти поля теперь есть в модели
                    responsible_person=row['responsible'],
                    work_type=row['work_type'],
                    measure_unit=row['measure_unit']
                )
                db.add(branch_object)
                db.commit()

                results.append({
                    "branch_code": branch_code,
                    "status": "success",
                    "object_id": branch_object.id
                })

            except Exception as e:
                results.append({
                    "branch_code": branch_code,
                    "status": "error",
                    "error": str(e)
                })
                continue

        return {"results": results}

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка обработки файла: {str(e)}")

@app.get("/reset-db", include_in_schema=False)
def reset_db():
    """Опасная функция - полностью пересоздает структуру БД!"""
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    return {"message": "Database structure recreated"}

@app.get("/api/branches/{branch_code}/territory", response_model=dict)
def get_branch_territory(branch_code: str, db: Session = Depends(get_db)):
    branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Филиал не найден")

    territory = db.query(BranchTerritory).filter(
        BranchTerritory.branch_id == branch.id
    ).first()

    if not territory:
        raise HTTPException(status_code=404, detail="Данные о территории не найдены")

    return {
        "branch_code": branch_code,
        "responsible_person": territory.responsible_person,
        "legal_area": territory.legal_area,
        "summer_area": territory.summer_area,
        "winter_area": territory.winter_area,
        "municipal_area": territory.municipal_area,
        "client_area": territory.client_area
    }

@app.post("/api/branches/{branch_code}/attachments/", response_model=AttachmentResponse)
async def upload_file_to_branch(
        branch_code: str,
        file_type: str = Form("other"),
        object_id: Optional[int] = Form(None),
        file: UploadFile = File(...),
        db: Session = Depends(get_db)
):
    """
    Загружает файл и привязывает его к филиалу

    Параметры:
    - branch_code: Код филиала (ВСП)
    - file_type: Тип файла (photo, document, plan)
    - object_id: ID объекта (опционально)
    - file: Загружаемый файл
    """
    # Проверяем существование филиала
    branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
    if not branch:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Филиал с кодом {branch_code} не найден"
        )

    # Если указан object_id, проверяем существование объекта
    if object_id:
        obj = db.query(BranchObject).filter(
            BranchObject.id == object_id,
            BranchObject.branch_id == branch.id
        ).first()
        if not obj:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Объект с ID {object_id} не найден в филиале {branch_code}"
            )

    # Генерируем уникальное имя файла
    file_ext = Path(file.filename).suffix
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)

    # Сохраняем файл на диск
    try:
        with open(file_path, "wb") as buffer:
            buffer.write(await file.read())
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ошибка при сохранении файла: {str(e)}"
        )

    # Сохраняем информацию о файле в БД
    db_attachment = BranchAttachment(
        branch_id=branch.id,
        object_id=object_id,
        file_type=file_type,
        file_url=unique_filename,
        original_filename=file.filename
    )

    db.add(db_attachment)
    db.commit()
    db.refresh(db_attachment)

    return db_attachment

@app.get("/api/branches/{branch_code}/attachments/{attachment_id}/download")
async def download_branch_attachment(
        branch_code: str,
        attachment_id: int,
        preview: bool = False,
        db: Session = Depends(get_db)
):
    try:
        # Проверяем существование филиала
        branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
        if not branch:
            raise HTTPException(status_code=404, detail="Филиал не найден")

        # Проверяем существование вложения
        attachment = db.query(BranchAttachment).filter(
            BranchAttachment.id == attachment_id,
            BranchAttachment.branch_id == branch.id
        ).first()

        if not attachment:
            raise HTTPException(status_code=404, detail="Вложение не найдено")

        # Формируем путь к файлу
        file_path = os.path.join(UPLOAD_DIR, attachment.file_url)

        # Проверяем существование файла
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Файл не найден на сервере")

        # Определяем MIME-тип
        file_ext = Path(attachment.file_url).suffix.lower()
        media_type = {
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
            '.webp': 'image/webp',
            '.bmp': 'image/bmp',
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.xls': 'application/vnd.ms-excel',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }.get(file_ext, 'application/octet-stream')

        # Для предпросмотра изображений возвращаем изображение напрямую
        if preview and media_type.startswith('image/'):
            return FileResponse(file_path, media_type=media_type)

        # Кодируем имя файла для Content-Disposition
        from urllib.parse import quote
        filename = quote(attachment.original_filename)

        # Для скачивания возвращаем файл с заголовком Content-Disposition
        async with aiofiles.open(file_path, mode='rb') as file:
            contents = await file.read()

        return Response(
            content=contents,
            media_type=media_type,
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"
            }
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при скачивании файла: {str(e)}"
        )



# Эндпоинт для получения списка файлов филиала
@app.get("/api/branches/{branch_code}/attachments/",
         response_model=List[AttachmentResponse])
def get_branch_attachments(
        branch_code: str,
        object_id: Optional[int] = None,
        file_type: Optional[str] = None,
        db: Session = Depends(get_db)
):
    """
    Получает список всех файлов, привязанных к филиалу.

    Параметры:
    - branch_code: internal_code филиала (ВСП)
    - object_id: фильтр по ID объекта (опционально)
    - file_type: фильтр по типу файла (опционально)
    """
    branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
    if not branch:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Филиал с кодом {branch_code} не найден"
        )

    query = db.query(BranchAttachment).filter(
        BranchAttachment.branch_id == branch.id
    )

    if object_id is not None:
        query = query.filter(BranchAttachment.object_id == object_id)

    if file_type is not None:
        query = query.filter(BranchAttachment.file_type == file_type)

    attachments = query.all()

    return [{
        "id": a.id,
        "branch_code": branch_code,
        "object_id": a.object_id,
        "file_type": a.file_type,
        "file_url": f"/api/branches/{branch_code}/attachments/{a.id}/download",
        "original_filename": a.original_filename,
        "uploaded_at": a.uploaded_at
    } for a in attachments]

@app.get("/file-upload", include_in_schema=False)
async def file_upload_page(request: Request):
    return templates.TemplateResponse("file_upload.html", {"request": request})

# Эндпоинт для удаления файла
@app.delete("/api/branches/{branch_code}/attachments/{attachment_id}/", status_code=status.HTTP_204_NO_CONTENT)
async def delete_branch_attachment(
        branch_code: str,
        attachment_id: int,
        db: Session = Depends(get_db)
):
    """
    Удаляет файл филиала

    Параметры:
    - branch_code: Код филиала (ВСП)
    - attachment_id: ID вложения
    """
    # Проверяем существование филиала
    branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
    if not branch:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Филиал с кодом {branch_code} не найден"
        )

    # Проверяем существование вложения
    attachment = db.query(BranchAttachment).filter(
        BranchAttachment.id == attachment_id,
        BranchAttachment.branch_id == branch.id
    ).first()

    if not attachment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Вложение не найдено"
        )

    # Удаляем файл с диска
    file_path = os.path.join(UPLOAD_DIR, attachment.file_url)
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ошибка при удалении файла: {str(e)}"
        )

    # Удаляем запись из БД
    db.delete(attachment)
    db.commit()

    return Response(status_code=status.HTTP_204_NO_CONTENT)




@app.post("/api/completed-works", response_model=CompletedWorkResponse)
def create_completed_work(work: CompletedWorkCreate, db: Session = Depends(get_db)):
    # Находим филиал по коду
    branch = db.query(Branch).filter(Branch.internal_code == work.branch_code).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Филиал с указанным кодом не найден")

    # Проверяем объект, если указан
    if work.object_id:
        if not db.query(BranchObject).filter(BranchObject.id == work.object_id).first():
            raise HTTPException(status_code=404, detail="Объект филиала не найден")

    # Здесь work.completion_date уже преобразован в datetime валидатором
    db_work = CompletedWork(
        branch_id=branch.id,
        object_id=work.object_id,
        work_type=work.work_type,
        completion_date=work.completion_date,
        responsible_person=work.responsible_person,
        notes=work.notes
    )

    db.add(db_work)
    db.commit()
    db.refresh(db_work)

    return {
        "branch_code": work.branch_code,
        "object_id": db_work.object_id,
        "work_type": db_work.work_type,
        "completion_date": db_work.completion_date,
        "responsible_person": db_work.responsible_person,
        "notes": db_work.notes
    }

@app.post("/api/attachments", response_model=AttachmentCreate)
async def create_attachment(
    attachment: AttachmentCreate,
    db: Session = Depends(get_db)
):
    # 1. Находим филиал по коду
    branch = db.query(Branch).filter(Branch.internal_code == attachment.branch_code).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Филиал не найден")

    # 2. Проверяем объект, если указан
    if attachment.object_id:
        obj = db.query(BranchObject).filter(
            BranchObject.id == attachment.object_id,
            BranchObject.branch_id == branch.id
        ).first()
        if not obj:
            raise HTTPException(status_code=404, detail="Объект не найден")

    # 3. Сохраняем в базу
    db_attachment = BranchAttachment(
        branch_id=branch.id,
        object_id=attachment.object_id,
        file_type=attachment.file_type,
        file_url=str(attachment.file_url)  # Преобразуем HttpUrl в строку
    )

    db.add(db_attachment)
    db.commit()
    db.refresh(db_attachment)

    return {
        "branch_code": attachment.branch_code,
        "object_id": attachment.object_id,
        "file_type": attachment.file_type,
        "file_url": attachment.file_url
    }

# Эндпоинты для чат-бота
@app.get("/api/branches", response_model=List[BranchCreate])
def search_branches(search: str = "", db: Session = Depends(get_db)):
    branches = db.query(Branch).filter(
        (Branch.address.ilike(f"%{search}%")) |
        (Branch.internal_code.ilike(f"%{search}%"))
    ).all()

    # Преобразуем SQLAlchemy объекты в словари
    return [
        {
            "address": branch.address,
            "internal_code": branch.internal_code,
            "latitude": branch.latitude,
            "longitude": branch.longitude
        }
        for branch in branches
    ]


@app.get("/api/branches/by-code/{branch_code}/objects", response_model=List[ObjectResponse])
def get_branch_objects(branch_code: str, db: Session = Depends(get_db)):
    try:
        branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
        if not branch:
            raise HTTPException(status_code=404, detail="Филиал не найден")

        objects = db.query(BranchObject).filter(BranchObject.branch_id == branch.id).all()

        result = []
        for obj in objects:
            obj_type = db.query(ObjectType).filter(ObjectType.id == obj.object_type_id).first()
            result.append({
                "id": obj.id,
                "branch_code": branch_code,
                "address": branch.address,
                "responsible": obj.responsible_person or "-",
                "zone": obj.zone or "-",
                "surface_type": obj.surface_type or "-",
                "work_type": obj.work_type or "-",
                "measure_unit": obj.measure_unit or (obj_type.measure_unit if obj_type else "-"),
                "work_volume": obj.area or "-",
                "work_period": obj.work_period or "-"
            })

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка сервера: {str(e)}")

@app.get("/api/object-types", response_model=List[ObjectTypeResponse])
def get_object_types(search: str = "", db: Session = Depends(get_db)):
    """Получение всех типов объектов с возможностью поиска"""
    query = db.query(ObjectType)
    if search:
        query = query.filter(ObjectType.name.ilike(f"%{search}%"))
    return query.all()


@app.get("/api/branches/by-code/{branch_code}/plans", response_model=List[MaintenancePlanResponse])
def get_branch_plans_by_code(branch_code: str, db: Session = Depends(get_db)):
    # Находим филиал по коду
    branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Филиал не найден")

    # Получаем планы для найденного филиала
    plans = db.query(MaintenancePlan).filter(MaintenancePlan.branch_id == branch.id).all()

    return [
        {
            "branch_code": branch_code,  # Возвращаем код вместо ID
            "object_id": plan.object_id,
            "work_type": plan.work_type,
            "frequency": plan.frequency,
            "next_maintenance_date": plan.next_maintenance_date
        }
        for plan in plans
    ]


@app.get("/api/branches/by-code/{branch_code}/completed-works", response_model=List[CompletedWorkResponse])
def get_branch_completed_works_by_code(branch_code: str, db: Session = Depends(get_db)):
    # Находим филиал по коду
    branch = db.query(Branch).filter(Branch.internal_code == branch_code).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Филиал не найден")

    # Получаем выполненные работы для филиала
    works = db.query(CompletedWork).filter(CompletedWork.branch_id == branch.id).all()

    return [
        {
            "branch_code": branch_code,  # Возвращаем код вместо ID
            "object_id": work.object_id,
            "work_type": work.work_type,
            "completion_date": work.completion_date,
            "responsible_person": work.responsible_person,
            "notes": work.notes
        }
        for work in works
    ]



@app.get("/api/nlp-query")
def process_nlp_query(query: str, db: Session = Depends(get_db)):
    if "планируется" in query.lower() or "планы" in query.lower():
        branch_code = "".join([c for c in query if c.isdigit()])
        if branch_code:
            try:
                plans = get_branch_plans(int(branch_code), db)
                return {
                    "status": "success",
                    "data": plans,
                    "message": f"Найдены планы для филиала {branch_code}"
                }
            except Exception as e:
                return {
                    "status": "error",
                    "message": f"Ошибка при получении планов: {str(e)}"
                }

    elif "выполнено" in query.lower() or "сделано" in query.lower():
        branch_code = "".join([c for c in query if c.isdigit()])
        if branch_code:
            try:
                works = get_branch_completed_works(int(branch_code), db)
                return {
                    "status": "success",
                    "data": works,
                    "message": f"Найдены выполненные работы для филиала {branch_code}"
                }
            except Exception as e:
                return {
                    "status": "error",
                    "message": f"Ошибка при получении выполненных работ: {str(e)}"
                }

    return {
        "status": "not_found",
        "message": "Не удалось обработать запрос. Уточните параметры поиска."
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=9080)
